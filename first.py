import threading
import time
from time import sleep
import json
import random
from datetime import *
from urllib import request
import sys
import shelve
import requests
import zipfile 
import openpyxl
import os
import itertools
import asyncio
import redis
import jieba
import jieba.analyse
from tqdm import tqdm
import configparser
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from html.parser import HTMLParser
# from dateutil.relativedelta import relativedelta


class test(threading.Thread):
    def __init__(self,num):
        threading.Thread.__init__(self)
        self.num = num

    def run(self):
        global count,mutex
        threadname = threading.currentThread().getName()
        for x in range(0,int(self.num)):
            mutex.acquire()
            count = count + 1
            mutex.release()
            #time.sleep(1)
        # print (threadname,self.num,count)

def yieldTest(key):
    key = yield key
    while True:
        key = key * 100
        key = yield key

async def compute(x, y):
    print("Compute %s + %s ..." % (x, y))
    await asyncio.sleep(3.0)
    return x + y
 
async def print_sum(x, y):
    result = await compute(x, y)
    print("%s + %s = %s" % (x, y, result))
       
if __name__ == "__main__":
##########threading#################
    global count,mutex
    pool = []
    count = 0
    mutex = threading.Lock()

    for x in range(0,10):
        pool.append(test(10))
    
    for t in pool:
        t.start()
    
    for t in pool:
        t.join()
###########json############################
    data = dict(name='Thread',value=count)
    dataJson = json.dumps(data)
    dataEpain = json.loads(dataJson)
    #print(dataEpain)
###########date#################################
    date = datetime.now().strftime('%Y-%m-01')
    tomorrow = (datetime.now()+timedelta(days=1)).strftime('%Y-%m-%d')
    week = (datetime.now()+timedelta(days=1)).weekday()
    day = datetime.now().day
    daytmp = '5'
    
    # dat = time.strftime("%Y-%m-%d",time.strptime("2009-08-08", "%Y-%m-%d"))
    # t = time.strptime("2009-08-08", "%Y-%m-%d")
    # y,m,d = t[0:3]
    print(datetime.now().date())
    print(type(datetime.strptime("2009-08-08", "%Y-%m-%d")))
    print((datetime(2006,10,12)-(datetime(2006,10,12)-timedelta(hours=24,seconds=30))).seconds)
    # if day==int(daytmp):
    #     print('true')
    # else:
    #     print('false')
    # nextmonth = (datetime.now() + relativedelta(months=1)).strftime('%Y-%m-%d')
    # print(date,tomorrow,week,day)
###########urllib curl访问#################################
req = request.Request('http://www.liaoxuefeng.com/files/attachments/00146917760141137418dc990d3459d99ea875458da0ea4000/0')
req.add_header('User-Agent', 'Mozilla/6.0 (iPhone; CPU iPhone OS 8_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/8.0 Mobile/10A5376e Safari/8536.25')
with request.urlopen(req) as f:
    img = open('downloadtest.jpg','wb')
    img.write(f.read())
    img.close()
    # print('Status:', f.status, f.reason)
    # for k, v in f.getheaders():
    #     print('%s: %s' % (k, v))
    # print('Data:', f.read().decode('utf-8'))

#############字符串format######################################
str1 = '{domain}, {year}'.format(domain='www.pythontab.com', year=2016)
str1 = '{:,}'.format(2016072145646)
str1 = '%s,%s'%(1,2)
arr = [1,2,3,4,5]
arrStr = ','.join(map(str, arr))
#print(arrStr)
arrRamdom = ['haha','heihei','hehe']
# print(random.choice(arrRamdom))
####################################################
dict_test = {}
dict_test['name'] = 'aaa'
dict_test['age'] = 12
sex = dict_test.get('sex',0)
# print(dict(zip(['a','b','c'],[1,2,3])))
# print(sex)
###################################################
# with open("txt.txt") as file:
#     data = file.read()
    ##print(data)
###################
day = datetime.now()
aaaaa = "select p_id, avg(rank) from amazon.ranking_top_{0:%Y%m} where rank > 0 and pc_id = %s and rank_time >= %s and rank_time < %s group by p_id having avg(rank) <= 10000 order by avg(rank) limit 10000".format(day)
# print(aaaaa)
########yield##########
x = 1
y = 2
r = "ok" if x>2 else "not ok"
# print(tuple("sales_index_{}".format(d) for d in range(1, 32)))
g = yieldTest(1)
print(g.send(None))
# print(g.__next__())
# print(g.__next__())
print(g.send(5))
print(g.send(6))
print(g.send(9))
# print(g.__next__())
###获取脚本参数
print(sys.argv)
########shelve
s = shelve.open("test.db")
s["contents"] = {"first":"base knowledge","second":"day day up"}
s.close()

with shelve.open("test.db", writeback=True) as f:
    print(f["contents"])
#########requests
html = requests.get("http://www.rakuten.co.jp/1bankanwebshop/")
print(html.cookies)
####################file split
imgfile = '01.jpg'
size = 1024*1024
if not os.path.exists('newdir'):
    os.mkdir("newdir")
with open(imgfile,'rb') as f:
    part = 1
    while True:
        content = f.read(size)
        if not content:
            break
        fileName = "newdir/imgpart_{0}".format(part)
        with open(fileName,'wb') as wf:
            wf.write(content)
        part += 1
# os.remove('imgpart_1')
# os.remove('imgpart_2')
###################file combine
files = os.listdir('newdir')
os.chdir('newdir')
os.remove('01_copy.jpg')
with open('01_copy.jpg','wb+') as wf:
    for f in files:
        if f.endswith('.zip'):
            continue
        with open(f,'rb') as inf:
            wf.write(inf.read())
##################zipfile
z = zipfile.ZipFile('01_copy.zip', 'w')
z.write('imgpart_1')
z.write('imgpart_2')
z.close()
###############excal
os.chdir('../')
dest_filename = 'empty_book.xlsx'
#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('empty_book.xlsx')
#ws1 = wb.active
ws1 = wb.create_sheet()
ws1.title = "range names2"
rows = [
    ['一行', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
for row in rows:
    ws1.append(row)
wb.save(filename = dest_filename)
###############itertools.groupby
fixId = [1493, 1493, 1493, 1493, 1493, 1493, 1493, 1493, 1493, 1493, 1493, 2884, 2884, 2884, 2884, 1493, 1493]
fixId = itertools.groupby(fixId,lambda x:x>2000)
for k,v in fixId:
    print(k,list(v))
################asyncio
loop = asyncio.get_event_loop()
tasks = [print_sum(1,2),print_sum(3,4)]
loop.run_until_complete(asyncio.wait(tasks))
loop.close()
################tqdm
dw_config = {'host':'127.0.0.1', 'user':'', 'password':'', 'db':'', 'port': 1, 'charset':'utf8'}
for i in tqdm(dw_config.items()):
    # print(i)
    sleep(0.01)
# for char in tqdm(["a", "b", "c", "d"]):
#     print(char)
##################redis
# r = redis.StrictRedis(host='localhost', port=6379, db=0)
# pool = redis.ConnectionPool(host='localhost', port=6379, db=0)
# r = redis.Redis(connection_pool=pool)
# r.set('bing', 'baz')
# pipe = r.pipeline()
# pipe.set('foo', 'bar')
# pipe.get('bing')
# result = pipe.execute()
# # print(r.get('bing').decode('UTF-8'))
# r.mset(dw_config)
# print(r.get('host').decode('UTF-8'))
# dic={"a1":"aa","b1":"bb"}
# r.hmset("dic_name",dic)
# print(r.hget("dic_name","b1"))
# li=["a1","b1"]
# print(r.hmget("dic_name",li))
#####################jieba
seg_list = jieba.analyse.extract_tags("13年左右多美滋金装优阶优创力优3段900克+100克1000克听装1-3岁")
print("Default Mode: " + "/ ".join(seg_list))
#####################config
# cf = configparser.ConfigParser()
# cf.read("my.conf")
# cf.get("mysql","db_host")
# o1 = cf.items("mysql")
# print(o1)
###################selenium
# driver = webdriver.Chrome() 
driver = webdriver.PhantomJS()
driver.get("http://www.baidu.com")
# assert "Python" in driver.title
elem = driver.find_element_by_name("wd")
elem.send_keys("dota2")
elem.send_keys(Keys.RETURN)
#driver.find_element_by_xpath("html/body/div[1]/div[2]/a[2]").click()
# driver.find_element_by_xpath("//*[@id='s_tab']/a[2]").click()
# driver.find_element_by_link_text("贴吧").click()
# elem = driver.find_element_by_id("s_tab")
# driver.get("https://www.google.co.in/")
# all_links = elem.find_elements_by_tag_name("a")
# help(all_links[1])
# for a in driver.find_elements_by_css_selector("div.c-container"):
#     print(a.get_attribute('id'))
for a in driver.find_elements_by_xpath("//*[@id='s_tab']/a"):
    print(a.get_attribute('href'))
# print(driver.page_source)
driver.quit()
#########################HTMLParser
print(HTMLParser().unescape('&#12468;&#12540;&#12523;&#12487;&#12531;&#12505;&#12450;)goldenbear(&#12468;&#12540;&#12523;&#12487;&#12531;&#12505;&#12450;) &#12459;&#12483;&#12488;&#12477;&#12540;'))
